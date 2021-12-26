# UniExcel: United Excels

import openpyxl
import openpyxl.styles
import sys

# Optional
try:
    from msoffcrypto import OfficeFile  # pip install msoffcrypto-tool
except ImportError:
    sys.stderr.write("[UniExcel.Warning] msoffcrypto not installed. Support of opening password protected worksheets is disabled. (Use `pip install msoffcrypto-tool` to install)\n")
    OfficeFile = None


VERSION = "UniExcel 2.1 (Build 20200319.1) Stable"


class UniExcel(object):
    @staticmethod
    def read_excel(ws, skip_title, use_legacy, debug_print=False):
        excel_data = []
        if use_legacy:
            for i in range(0, skip_title):
                excel_data.append([None] * ws.max_column)

        cnt_row = 0
        for row_data in ws.iter_rows(min_row=skip_title + 1, values_only=True):
            cnt_row += 1
            this_line = list(row_data)
            if debug_print:
                print("<{}> {}".format(cnt_row, this_line))
            excel_data.append(this_line)
        return excel_data

    @staticmethod
    def solid_fill(color_rgb):
        return openpyxl.styles.PatternFill("solid", fgColor=color_rgb)  # "FFFF00"

    # If use_legacy is set to True, pre-loaded values will index from `skip_title` (instead of 0).
    # This should only be used for backward compatibility, and might have several issues.
    def __init__(self, filepath=None, read_all=True, skip_title=1, default_worksheet_index=0, use_legacy=False, password=None, tmpfile=None):
        self.max_row = 0
        self.max_column = 0
        self.excel_data = []
        self.wb = None
        self.ws = None

        if password:
            if not OfficeFile:
                raise NotImplementedError("msoffcrypto-tool is required.")
            if not tmpfile:
                raise Exception("Parameter `tmpfile` must be provided in order to open encrypted excel.")
            if not filepath:
                raise Exception("Parameter `filepath` must be provided in order to open encrypted excel.")
            try:
                enc_excel = OfficeFile(open(filepath, "rb"))
                enc_excel.load_key(password=password)
                enc_excel.decrypt(open(tmpfile, "wb"))
            except Exception:
                print("[UniExcel.Error] Unable to decrypt excel.")
                raise
            filepath = tmpfile

        if filepath:
            self.wb = openpyxl.load_workbook(filepath)
            self.use_sheet(default_worksheet_index, read_all=read_all, skip_title=skip_title, use_legacy=use_legacy)
        else:
            self.wb = openpyxl.Workbook()
            self.use_sheet(0, read_all=False)

    def use_sheet(self, index, read_all=True, skip_title=1, use_legacy=False):
        self.ws = self.wb.worksheets[index]
        if read_all:
            self.max_row = self.ws.max_row
            self.max_column = self.ws.max_column
            self.excel_data = UniExcel.read_excel(self.ws, skip_title, use_legacy)
        else:
            self.max_row = 0
            self.max_column = 0
            self.excel_data = []

    def set_title(self, title, index=None):
        if index is None:
            self.ws.title = title
        else:
            self.wb.worksheets[index].title = title

    def create_sheet(self, title, index=None):
        self.wb.create_sheet(title, index=index)
        self.ws = self.wb[title]
        self.max_row = 0
        self.max_column = 0
        self.excel_data = []

        # find index of this sheet...
        for index in range(len(self.wb.worksheets)):
            if self.wb.worksheets[index].title == title:
                return index

    def get_value(self, row, col):
        return self.ws.cell(row=row, column=col).value

    def set_value(self, row, col, value):
        self.ws.cell(row=row, column=col).value = value

    def new_cell(self, value):
        cell = openpyxl.cell.WriteOnlyCell(self.ws, value=value)
        return cell

    def append_cells(self, cells):
        self.ws.append(cells)

    def append(self, values, fill=None):
        if fill:
            cells = [self.new_cell(x) for x in values]
            for cell in cells:
                cell.fill = fill
            self.append_cells(cells)
        else:
            self.ws.append(values)

    def save(self, filepath):
        self.wb.save(filepath)
